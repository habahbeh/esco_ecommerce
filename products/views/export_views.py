# File: products/views/export_views.py
"""
Export views for products data
Handles CSV, Excel, and JSON exports with proper permissions
"""

from typing import Dict, Any, List, Optional
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.views.decorators.http import require_http_methods
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.utils.translation import gettext as _
from django.utils import timezone
from django.db.models import Q, Avg, Count, Sum, F
from django.views import View
import csv
import json
import logging
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .base_views import AdminRequiredMixin
from ..models import Product, Category, Brand, ProductReview, Tag

logger = logging.getLogger(__name__)


class ExportMixin:
    """Mixin for export functionality"""

    def check_export_permissions(self, request):
        """Check if user has export permissions"""
        if not request.user.is_staff:
            raise PermissionError(_('غير مصرح لك بتصدير البيانات'))

    def get_filename(self, base_name: str, file_format: str) -> str:
        """Generate filename with timestamp"""
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        return f"{base_name}_{timestamp}.{file_format}"

    def prepare_csv_response(self, filename: str) -> HttpResponse:
        """Prepare CSV HTTP response"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write('\ufeff'.encode('utf8'))  # BOM for Excel compatibility
        return response

    def prepare_excel_response(self, filename: str) -> HttpResponse:
        """Prepare Excel HTTP response"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def prepare_json_response(self, filename: str) -> HttpResponse:
        """Prepare JSON HTTP response"""
        response = HttpResponse(content_type='application/json; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


@method_decorator(staff_member_required, name='dispatch')
class ExportProductsView(View, ExportMixin):
    """
    Export products data in various formats
    """

    def get(self, request):
        """Handle export request"""
        try:
            self.check_export_permissions(request)

            export_format = request.GET.get('format', 'csv').lower()
            export_type = request.GET.get('type', 'all')  # all, filtered, selected

            # Get products based on export type
            products = self.get_products_queryset(request, export_type)

            if export_format == 'csv':
                return self.export_csv(products)
            elif export_format == 'excel':
                return self.export_excel(products)
            elif export_format == 'json':
                return self.export_json(products)
            else:
                return JsonResponse({
                    'error': _('صيغة التصدير غير مدعومة')
                }, status=400)

        except PermissionError as e:
            return HttpResponseForbidden(str(e))
        except Exception as e:
            logger.error(f"Error in products export: {e}")
            return JsonResponse({
                'error': _('حدث خطأ أثناء التصدير')
            }, status=500)

    def get_products_queryset(self, request, export_type: str):
        """Get products queryset based on export type"""
        base_queryset = Product.objects.select_related(
            'category', 'brand', 'created_by'
        ).prefetch_related('tags')

        if export_type == 'selected':
            # Get selected product IDs from request
            product_ids = request.GET.getlist('ids')
            if product_ids:
                return base_queryset.filter(id__in=product_ids)
            else:
                return base_queryset.none()

        elif export_type == 'filtered':
            # Apply filters similar to product list view
            return self.apply_export_filters(base_queryset, request)

        # Default: export all products
        return base_queryset.all()

    def apply_export_filters(self, queryset, request):
        """Apply filters for export"""
        # Status filter
        status = request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        # Active filter
        is_active = request.GET.get('is_active')
        if is_active:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')

        # Category filter
        category_id = request.GET.get('category')
        if category_id:
            try:
                queryset = queryset.filter(category_id=int(category_id))
            except ValueError:
                pass

        # Brand filter
        brand_id = request.GET.get('brand')
        if brand_id:
            try:
                queryset = queryset.filter(brand_id=int(brand_id))
            except ValueError:
                pass

        # Date range filter
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')

        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(created_at__date__gte=date_from)
            except ValueError:
                pass

        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(created_at__date__lte=date_to)
            except ValueError:
                pass

        return queryset

    def export_csv(self, products):
        """Export products to CSV"""
        filename = self.get_filename('products_export', 'csv')
        response = self.prepare_csv_response(filename)

        writer = csv.writer(response)

        # Write header
        headers = [
            'ID', 'اسم المنتج', 'Product Name (EN)', 'SKU', 'الفئة', 'العلامة التجارية',
            'السعر الأساسي', 'نسبة الخصم', 'مبلغ الخصم', 'السعر الحالي', 'المخزون',
            'تتبع المخزون', 'حالة المخزون', 'الحد الأدنى للمخزون', 'الوزن', 'الأبعاد',
            'الحالة', 'نشط', 'مميز', 'جديد', 'عدد المبيعات', 'عدد المشاهدات',
            'الوسوم', 'تاريخ الإنشاء', 'تاريخ التحديث'
        ]
        writer.writerow(headers)

        # Write data
        for product in products:
            row = [
                product.id,
                product.name,
                product.name_en or '',
                product.sku,
                product.category.name if product.category else '',
                product.brand.name if product.brand else '',
                float(product.base_price),
                product.discount_percentage or 0,
                float(product.discount_amount or 0),
                float(product.current_price),
                product.stock_quantity or 0,
                'نعم' if product.track_inventory else 'لا',
                product.get_stock_status_display(),
                product.low_stock_threshold or 0,
                float(product.weight) if product.weight else 0,
                product.dimensions or '',
                product.get_status_display(),
                'نعم' if product.is_active else 'لا',
                'نعم' if product.is_featured else 'لا',
                'نعم' if product.is_new else 'لا',
                product.sales_count,
                product.views_count,
                ', '.join([tag.name for tag in product.tags.all()]),
                product.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                product.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            writer.writerow(row)

        # Log export
        logger.info(f"Admin {self.request.user.id} exported {products.count()} products to CSV")

        return response

    def export_excel(self, products):
        """Export products to Excel"""
        if not OPENPYXL_AVAILABLE:
            return JsonResponse({
                'error': _('مكتبة openpyxl غير متوفرة لتصدير Excel')
            }, status=400)

        filename = self.get_filename('products_export', 'xlsx')
        response = self.prepare_excel_response(filename)

        # Create workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Products"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            'ID', 'اسم المنتج', 'Product Name (EN)', 'SKU', 'الفئة', 'العلامة التجارية',
            'السعر الأساسي', 'نسبة الخصم', 'مبلغ الخصم', 'السعر الحالي', 'المخزون',
            'تتبع المخزون', 'حالة المخزون', 'الحد الأدنى للمخزون', 'الوزن', 'الأبعاد',
            'الحالة', 'نشط', 'مميز', 'جديد', 'عدد المبيعات', 'عدد المشاهدات',
            'الوسوم', 'تاريخ الإنشاء', 'تاريخ التحديث'
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_num, product in enumerate(products, 2):
            data = [
                product.id,
                product.name,
                product.name_en or '',
                product.sku,
                product.category.name if product.category else '',
                product.brand.name if product.brand else '',
                float(product.base_price),
                product.discount_percentage or 0,
                float(product.discount_amount or 0),
                float(product.current_price),
                product.stock_quantity or 0,
                'نعم' if product.track_inventory else 'لا',
                product.get_stock_status_display(),
                product.low_stock_threshold or 0,
                float(product.weight) if product.weight else 0,
                product.dimensions or '',
                product.get_status_display(),
                'نعم' if product.is_active else 'لا',
                'نعم' if product.is_featured else 'لا',
                'نعم' if product.is_new else 'لا',
                product.sales_count,
                product.views_count,
                ', '.join([tag.name for tag in product.tags.all()]),
                product.created_at,
                product.updated_at
            ]

            for col_num, value in enumerate(data, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to response
        workbook.save(response)

        # Log export
        logger.info(f"Admin {self.request.user.id} exported {products.count()} products to Excel")

        return response

    def export_json(self, products):
        """Export products to JSON"""
        filename = self.get_filename('products_export', 'json')
        response = self.prepare_json_response(filename)

        # Prepare data
        products_data = []
        for product in products:
            # Get additional data
            avg_rating = product.reviews.filter(is_approved=True).aggregate(
                avg=Avg('rating')
            )['avg']

            product_data = {
                'id': product.id,
                'name': product.name,
                'name_en': product.name_en,
                'slug': product.slug,
                'sku': product.sku,
                'description': product.description,
                'short_description': product.short_description,
                'category': {
                    'id': product.category.id if product.category else None,
                    'name': product.category.name if product.category else None,
                    'slug': product.category.slug if product.category else None
                },
                'brand': {
                    'id': product.brand.id if product.brand else None,
                    'name': product.brand.name if product.brand else None,
                    'slug': product.brand.slug if product.brand else None
                },
                'pricing': {
                    'base_price': float(product.base_price),
                    'discount_percentage': product.discount_percentage,
                    'discount_amount': float(product.discount_amount or 0),
                    'current_price': float(product.current_price),
                    'has_discount': product.has_discount
                },
                'inventory': {
                    'stock_quantity': product.stock_quantity,
                    'track_inventory': product.track_inventory,
                    'stock_status': product.stock_status,
                    'low_stock_threshold': product.low_stock_threshold,
                    'in_stock': product.in_stock
                },
                'attributes': {
                    'weight': float(product.weight) if product.weight else None,
                    'dimensions': product.dimensions,
                    'specifications': product.specifications
                },
                'status': {
                    'status': product.status,
                    'is_active': product.is_active,
                    'is_featured': product.is_featured,
                    'is_new': product.is_new
                },
                'statistics': {
                    'sales_count': product.sales_count,
                    'views_count': product.views_count,
                    'average_rating': round(float(avg_rating), 2) if avg_rating else 0,
                    'review_count': product.reviews.filter(is_approved=True).count()
                },
                'tags': [tag.name for tag in product.tags.all()],
                'timestamps': {
                    'created_at': product.created_at.isoformat(),
                    'updated_at': product.updated_at.isoformat()
                }
            }
            products_data.append(product_data)

        # Export metadata
        export_data = {
            'metadata': {
                'export_date': timezone.now().isoformat(),
                'total_products': len(products_data),
                'exported_by': self.request.user.username,
                'format_version': '1.0'
            },
            'products': products_data
        }

        # Write JSON
        json.dump(export_data, response, ensure_ascii=False, indent=2, default=str)

        # Log export
        logger.info(f"Admin {self.request.user.id} exported {len(products_data)} products to JSON")

        return response


@method_decorator(staff_member_required, name='dispatch')
class ExportReviewsView(View, ExportMixin):
    """
    Export product reviews data
    """

    def get(self, request):
        """Handle reviews export request"""
        try:
            self.check_export_permissions(request)

            export_format = request.GET.get('format', 'csv').lower()

            # Get reviews queryset
            reviews = self.get_reviews_queryset(request)

            if export_format == 'csv':
                return self.export_reviews_csv(reviews)
            elif export_format == 'excel':
                return self.export_reviews_excel(reviews)
            else:
                return JsonResponse({
                    'error': _('صيغة التصدير غير مدعومة')
                }, status=400)

        except PermissionError as e:
            return HttpResponseForbidden(str(e))
        except Exception as e:
            logger.error(f"Error in reviews export: {e}")
            return JsonResponse({
                'error': _('حدث خطأ أثناء التصدير')
            }, status=500)

    def get_reviews_queryset(self, request):
        """Get reviews queryset with filters"""
        queryset = ProductReview.objects.select_related(
            'product', 'user'
        ).order_by('-created_at')

        # Apply filters
        status = request.GET.get('status')
        if status == 'approved':
            queryset = queryset.filter(is_approved=True)
        elif status == 'pending':
            queryset = queryset.filter(is_approved=False)

        rating = request.GET.get('rating')
        if rating:
            try:
                queryset = queryset.filter(rating=int(rating))
            except ValueError:
                pass

        product_id = request.GET.get('product_id')
        if product_id:
            try:
                queryset = queryset.filter(product_id=int(product_id))
            except ValueError:
                pass

        return queryset

    def export_reviews_csv(self, reviews):
        """Export reviews to CSV"""
        filename = self.get_filename('reviews_export', 'csv')
        response = self.prepare_csv_response(filename)

        writer = csv.writer(response)

        # Write header
        headers = [
            'ID', 'المنتج', 'المستخدم', 'التقييم', 'العنوان', 'المحتوى',
            'مفيد', 'غير مفيد', 'البلاغات', 'معتمد', 'تاريخ الإنشاء'
        ]
        writer.writerow(headers)

        # Write data
        for review in reviews:
            row = [
                review.id,
                review.product.name,
                review.user.username if review.user else 'مجهول',
                review.rating,
                review.title or '',
                review.content,
                review.helpful_count,
                review.not_helpful_count,
                review.report_count,
                'نعم' if review.is_approved else 'لا',
                review.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            writer.writerow(row)

        # Log export
        logger.info(f"Admin {self.request.user.id} exported {reviews.count()} reviews to CSV")

        return response

    def export_reviews_excel(self, reviews):
        """Export reviews to Excel"""
        if not OPENPYXL_AVAILABLE:
            return JsonResponse({
                'error': _('مكتبة openpyxl غير متوفرة لتصدير Excel')
            }, status=400)

        filename = self.get_filename('reviews_export', 'xlsx')
        response = self.prepare_excel_response(filename)

        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Reviews"

        # Apply styles (similar to products export)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            'ID', 'المنتج', 'المستخدم', 'التقييم', 'العنوان', 'المحتوى',
            'مفيد', 'غير مفيد', 'البلاغات', 'معتمد', 'تاريخ الإنشاء'
        ]

        # Write headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_num, review in enumerate(reviews, 2):
            data = [
                review.id,
                review.product.name,
                review.user.username if review.user else 'مجهول',
                review.rating,
                review.title or '',
                review.content,
                review.helpful_count,
                review.not_helpful_count,
                review.report_count,
                'نعم' if review.is_approved else 'لا',
                review.created_at
            ]

            for col_num, value in enumerate(data, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(response)

        # Log export
        logger.info(f"Admin {self.request.user.id} exported {reviews.count()} reviews to Excel")

        return response


# Legacy function-based views for backward compatibility
@staff_member_required
@require_http_methods(["GET"])
def export_products(request):
    """Legacy function view for products export"""
    view = ExportProductsView()
    view.request = request
    return view.get(request)


@staff_member_required
@require_http_methods(["GET"])
def export_reviews(request):
    """Legacy function view for reviews export"""
    view = ExportReviewsView()
    view.request = request
    return view.get(request)